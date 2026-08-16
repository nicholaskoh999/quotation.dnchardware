#!/usr/bin/env python3
"""Build pricing-engine-v2-input.xlsx — the blank workbook the business fills in.

Deliberately empty of business values. Every sheet carries its columns, a note
on what each column means, and where a shape genuinely needs demonstrating, one
row marked EXAMPLE — DELETE containing an obviously fake number. No rate, charge,
tier, markup or supplier cost in this file is real, and none of it is read by the
application.

    python3 tests/tools/build-pricing-workbook.py <output.xlsx>
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD_FILL = PatternFill('solid', fgColor='1C3FAA')
HEAD_FONT = Font(color='FFFFFF', bold=True, size=11)
NOTE_FONT = Font(color='505868', size=10, italic=True)
REQ_FILL = PatternFill('solid', fgColor='FFF7E6')
EX_FILL = PatternFill('solid', fgColor='FDECEC')
EX_FONT = Font(color='B42318', size=10)
TITLE_FONT = Font(bold=True, size=14, color='1C3FAA')
BODY_FONT = Font(size=11)
THIN = Side(style='thin', color='D6DAE3')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, purpose, columns, notes, examples=()):
    """columns: list of (header, required, meaning). examples: list of row lists."""
    ws = wb.create_sheet(name)
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = purpose
    ws['A2'].font = NOTE_FONT
    ws['A4'] = 'Columns marked REQUIRED must be filled in for the row to be usable.'
    ws['A4'].font = NOTE_FONT

    # meaning block, one line per column, above the table
    r = 6
    for head, required, meaning in columns:
        ws.cell(r, 1, head).font = Font(bold=True, size=10)
        ws.cell(r, 2, 'REQUIRED' if required else 'optional').font = Font(
            size=10, bold=required, color='B42318' if required else '505868')
        ws.cell(r, 3, meaning).font = NOTE_FONT
        r += 1
    r += 1

    for line in notes:
        ws.cell(r, 1, line).font = NOTE_FONT
        r += 1
    r += 1

    header_row = r
    for i, (head, required, _) in enumerate(columns, start=1):
        c = ws.cell(header_row, i, head)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    for j, row in enumerate(examples, start=header_row + 1):
        for i, val in enumerate(row, start=1):
            c = ws.cell(j, i, val)
            c.fill = EX_FILL
            c.font = EX_FONT
            c.border = BOX

    # empty rows, ready to type into
    first_blank = header_row + 1 + len(examples)
    for j in range(first_blank, first_blank + 60):
        for i in range(1, len(columns) + 1):
            c = ws.cell(j, i)
            c.border = BOX
            if columns[i - 1][1]:
                c.fill = REQ_FILL

    for i, (head, _, meaning) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(26, len(head) + 6))
    ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 30)
    return ws


def build(path):
    wb = Workbook()
    wb.remove(wb.active)

    # ── README ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet('README')
    ws.column_dimensions['A'].width = 118
    lines = [
        ('Pricing Engine V2 — input workbook', TITLE_FONT),
        ('', None),
        ('This workbook is EMPTY of business values on purpose. Nothing in it has been guessed.', NOTE_FONT),
        ('', None),
        ('WHAT THIS IS', Font(bold=True, size=12)),
        ('The current quotation system holds its rates inside the source code: a cost rate per finish, a', BODY_FONT),
        ('4140 QT table for two sizes, two thread-length bands, and a plating surcharge. They work, but the', BODY_FONT),
        ('people who own those numbers cannot see or change them. Pricing Engine V2 moves them into rules', BODY_FONT),
        ('the business maintains. This workbook is where those rules are written down for the first time.', BODY_FONT),
        ('', None),
        ('HOW TO FILL IT IN', Font(bold=True, size=12)),
        ('1. One sheet per kind of rule. Fill in the sheets you have real numbers for; leave the rest empty.', BODY_FONT),
        ('2. Columns shaded cream are REQUIRED — a row without them cannot be used and will be rejected.', BODY_FONT),
        ('3. Rows shaded pink and marked "EXAMPLE — DELETE" show the shape only. Their numbers are fake.', BODY_FONT),
        ('   Delete every one of them before sending the workbook back.', BODY_FONT),
        ('4. Every rule row needs an Effective Date. It is the date the rule STARTS applying, and a', BODY_FONT),
        ('   quotation is priced with the rules in force on the quotation date — not on today\'s date.', BODY_FONT),
        ('5. Never edit a rate in place once it has been used. Add a new row with a later Effective Date.', BODY_FONT),
        ('   The old row is what lets an old quotation still explain itself.', BODY_FONT),
        ('6. Cost Basis must say what the number is measured in: per_kg, per_pc, per_mm or percent.', BODY_FONT),
        ('   A charge whose basis is ambiguous is a wrong price waiting for a size it was not calibrated on.', BODY_FONT),
        ('7. Use the Remark column freely. "Agreed with supplier 12 Mar", "review after CNY" — the remark', BODY_FONT),
        ('   is what makes a rate reviewable a year later.', BODY_FONT),
        ('', None),
        ('THE TWO QUESTIONS THE AUDIT COULD NOT ANSWER', Font(bold=True, size=12)),
        ('· Supplier Cost Per KG vs Internal Cost Rate Per KG. Today one number does both jobs. If the rate', BODY_FONT),
        ('  you quote with is not the rate the supplier invoices you at, fill in BOTH columns. If they are', BODY_FONT),
        ('  the same number, say so in the Remark rather than leaving one blank.', BODY_FONT),
        ('· Whether a customer ever has a different COST, or only a different MARKUP. The design assumes', BODY_FONT),
        ('  cost is global and margin is per customer. If that is wrong for any customer, say which.', BODY_FONT),
        ('', None),
        ('WHAT HAPPENS NEXT', Font(bold=True, size=12)),
        ('Nothing changes on the live system when this workbook comes back. The rules are loaded read-only', BODY_FONT),
        ('first, then run in shadow mode beside the current calculation so every difference can be seen and', BODY_FONT),
        ('explained BEFORE any price moves. See pricing-engine-v2-plan.md, section 10.', BODY_FONT),
        ('', None),
        ('WHAT IS ALREADY IN THE CODE TODAY (for reference while filling this in)', Font(bold=True, size=12)),
        ('These are the values the application uses right now. They are stated here so you can decide what', NOTE_FONT),
        ('replaces them — they are NOT copied into the sheets, because a transcribed number with no owner', NOTE_FONT),
        ('is how they became invisible in the first place.', NOTE_FONT),
        ('  · MS sag rod cost rate per kg, by finish: PL 2.80, ZP 4.20, HDG 6.00', BODY_FONT),
        ('  · MS HDG additional cost: 1.60 per piece', BODY_FONT),
        ('  · 4140 QT sag rod, fullsize: M12 rate 8.50 add 2.50 · M16 rate 6.50 add 3.50', BODY_FONT),
        ('  · 4140 QT sag rod, undersize: M12 rate 9.50 add 2.00 · M16 rate 8.00 add 2.50', BODY_FONT),
        ('  · Finish surcharge on the rate: ZP +1.50, HDG +3.20', BODY_FONT),
        ('  · HDG thread brushing, on additional cost: +1.00', BODY_FONT),
        ('  · Thread length band, additional cost: TL 30-120mm -> 0.60 · TL over 120 to 200mm -> 1.00', BODY_FONT),
        ('  · No labour component, no quantity effect on unit cost, and no accessory markup exist today.', BODY_FONT),
        ('  · Unit weight = diameter(mm)^2 x developed length(mm) x 0.0000061654. V2 does not change it.', BODY_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(i, 1, text)
        if font:
            c.font = font

    # ── MATERIAL_RATES ──────────────────────────────────────────────────────
    sheet(wb, 'MATERIAL_RATES', 'MATERIAL_RATES',
          'What a kilogram of each material costs us, and what we quote it at.',
          [('Material', True, 'MS, S45C, S45C + HARDEN = G8.8, 4140 QT, 4140 QT + HARDEN = G10.9, 4140, 4340 QT, SS304, SS316, Y BAR'),
           ('Raw Material Type', False, 'Round bar, hex bar, plate — whatever distinguishes two stocks of the same grade'),
           ('Supplier', False, 'Who this rate comes from'),
           ('Size From', False, 'Smallest size this rate covers (M8, 1/2). Leave blank for all sizes'),
           ('Size To', False, 'Largest size this rate covers. Leave blank for all sizes'),
           ('Supplier Cost Per KG', False, 'What the supplier invoices us, before anything is added'),
           ('Internal Cost Rate Per KG', True, 'The rate the quotation calculates with. May differ from the supplier cost'),
           ('Effective Date', True, 'YYYY-MM-DD. The date this rate STARTS applying'),
           ('Active', True, 'Y or N. N keeps the row for history without using it'),
           ('Remark', False, 'Why this rate, and when to review it')],
          ['A size range is inclusive at both ends. Where two rows both match, the narrower range wins.',
           'If the internal rate equals the supplier cost, put the same number in both and say so in the Remark.'],
          [['MS', 'Round bar', 'EXAMPLE SUPPLIER', 'M8', 'M30', 0.00, 0.00, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zeroes, not real rates.']])

    # ── DIAMETER_RULES ──────────────────────────────────────────────────────
    sheet(wb, 'DIAMETER_RULES', 'DIAMETER_RULES',
          'The bar diameter each size is actually cut from — the number the weight is calculated on.',
          [('Product Type', True, 'SAG ROD, STUD, ANCHOR BOLT, U BOLT, SQ U BOLT, L BOLT, J BOLT, PLATE'),
           ('Material', True, 'As written in MATERIAL_RATES'),
           ('Size Type', True, 'FULLSIZE or UNDERSIZE. Blank for a product that has no size type (Stud)'),
           ('Size', True, 'M20, 1/2, 3/4 — as it is written on the quotation'),
           ('Diameter (mm)', True, 'The real bar diameter. This is what the weight is calculated on'),
           ('Effective Date', False, 'YYYY-MM-DD, if this changed on a date'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'Why this differs from the nominal size')],
          ['This sheet mirrors the Diameter Settings screen that is already live. Fill it in ONLY for',
           'sizes where the effective diameter differs from what the system already uses, or where you',
           'want the value recorded formally. The screen remains the source of truth either way.',
           'An UNDERSIZE M20 is not 20mm — the whole point of this table is that the difference is stated.'],
          [['SAG ROD', '4140 QT', 'UNDERSIZE', 'M12', 0.0, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zero is not a diameter.']])

    # ── PROCESS_COST_RULES ──────────────────────────────────────────────────
    sheet(wb, 'PROCESS_COST_RULES', 'PROCESS_COST_RULES',
          'What each manufacturing step adds to the cost of one piece.',
          [('Product Type', False, 'Blank means every product'),
           ('Material', False, 'Blank means every material'),
           ('Process', True, 'THREAD, BEND, CUTTING, MACHINING, FINISH, HEAT TREATMENT, OTHER'),
           ('Condition Field', False, 'The field the rule reads: TL, LENGTH, SIZE, QTY, WEIGHT. Blank means always'),
           ('Min Value', False, 'Lowest value of that field this rule covers, inclusive'),
           ('Max Value', False, 'Highest value, inclusive. Leave blank for no upper limit'),
           ('Cost Basis', True, 'per_pc, per_kg, per_mm or percent'),
           ('Cost', True, 'The number, in the basis above'),
           ('Effective Date', True, 'YYYY-MM-DD'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'What this charge covers')],
          ['Bands must MEET, not nearly meet: 30-120 followed by 121-200 leaves 120.5 uncharged. Write',
           '30-120 and then 120-200, and the later band wins the shared edge. (The live thread-length',
           'table had exactly this hole and it was closed in the previous audit pass.)',
           'One row per band. A rule with no condition field applies to every item that matches the',
           'product and material.'],
          [['L BOLT', '', 'BEND', 'SIZE', 'M8', 'M30', 'per_pc', 0.00, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zero is not a bending charge.']])

    # ── LABOUR_QTY_RULES ────────────────────────────────────────────────────
    sheet(wb, 'LABOUR_QTY_RULES', 'LABOUR_QTY_RULES',
          'How labour per piece changes with the size of the order.',
          [('Product Type', False, 'Blank means every product'),
           ('Material', False, 'Blank means every material'),
           ('Qty From', True, 'Smallest quantity in this band, inclusive'),
           ('Qty To', False, 'Largest quantity, inclusive. Blank means no upper limit'),
           ('Labour Cost Per PC', True, 'Labour added to ONE piece at this quantity'),
           ('Effective Date', True, 'YYYY-MM-DD'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'Setup time, minimum charge, whatever this band represents')],
          ['This is the first rule that would make quantity affect a unit price. Quantity still does NOT',
           'define what counts as the same item in Pricing History — a past order of 1 says nothing about',
           'what the item costs to make.',
           'Bands must cover every quantity from 1 upwards, or an order in the gap cannot be priced.'],
          [['', '', 1, 9, 0.00, '2026-01-01', 'N', 'EXAMPLE — DELETE. Zero is not a labour cost.']])

    # ── FINISH_COST_RULES ───────────────────────────────────────────────────
    sheet(wb, 'FINISH_COST_RULES', 'FINISH_COST_RULES',
          'What plating and galvanising cost, separately from the steel.',
          [('Finish', True, 'PL, ZP or HDG'),
           ('Material', False, 'Blank means every material'),
           ('Product Type', False, 'Blank means every product'),
           ('Size From', False, 'Smallest size this rule covers'),
           ('Size To', False, 'Largest size this rule covers'),
           ('Cost Basis', True, 'per_kg, per_pc or percent'),
           ('Cost', True, 'The number, in the basis above'),
           ('Effective Date', True, 'YYYY-MM-DD'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'Minimum lot charge, subcontractor, anything that qualifies it')],
          ['Today the finish is folded into the cost rate per kg (ZP +1.50, HDG +3.20) and partly into',
           'the additional cost (HDG thread brushing +1.00). Splitting it out is the point of this sheet:',
           'say what the finish itself costs, and on what basis.',
           'PL usually costs nothing. Enter it as 0 rather than leaving it out, so "no charge" is a',
           'recorded decision and not an omission.'],
          [['HDG', '', '', '', '', 'per_kg', 0.00, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zero is not a galvanising cost.']])

    # ── CUSTOMER_MARKUP ─────────────────────────────────────────────────────
    sheet(wb, 'CUSTOMER_MARKUP', 'CUSTOMER_MARKUP',
          'The margin each customer carries.',
          [('Customer', True, 'The company name exactly as it appears on the quotation'),
           ('Material', False, 'Blank means every material'),
           ('Product Type', False, 'Blank means every product'),
           ('Markup %', True, 'A percentage. 12 means 12%, not 0.12'),
           ('Effective Date', True, 'YYYY-MM-DD'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'Agreement, expiry, review date')],
          ['Most specific row wins: customer + product + material beats customer + material beats customer.',
           'A customer with no row here uses the general markup.',
           'The design assumes a customer changes the MARGIN, not the COST. If any customer genuinely has',
           'a different cost, do not put it here — say which customer and why, and it gets its own table.'],
          [['EXAMPLE CUSTOMER SDN BHD', '', '', 0, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zero is not a markup.']])

    # ── ACCESSORY_COSTS ─────────────────────────────────────────────────────
    sheet(wb, 'ACCESSORY_COSTS', 'ACCESSORY_COSTS',
          'What nuts and washers cost from the supplier, and at what margin we sell them.',
          [('Accessory Type', True, 'NUT, FLAT WASHER, SPRING WASHER, or the wording you use'),
           ('Material', True, 'MS, 4140 QT, SS304 — the accessory\'s own material, not the bolt\'s'),
           ('Finish', True, 'PL, ZP, HDG'),
           ('Size', True, 'M20, 1/2 — the accessory\'s size'),
           ('Supplier', False, 'Who supplies it'),
           ('Supplier Cost', True, 'What one piece costs us'),
           ('Markup %', True, 'The accessory\'s own margin. 12 means 12%'),
           ('Effective Date', True, 'YYYY-MM-DD'),
           ('Active', True, 'Y or N'),
           ('Remark', False, 'Pack size, minimum order, anything that qualifies the cost')],
          ['An accessory is a small product with its own cost, its own markup, its own supplier and its',
           'own history. It is NOT part of the bolt\'s unit price, and nothing here may be folded into one.',
           'A 4140 QT nut in PL at M20 is a different row from an MS nut in PL at M20. Same size, different',
           'item, different cost.'],
          [['NUT', 'MS', 'PL', 'M20', 'EXAMPLE SUPPLIER', 0.00, 0, '2026-01-01', 'N',
            'EXAMPLE — DELETE. Zero is not a supplier cost.']])

    wb.save(path)
    return path


if __name__ == '__main__':
    out = sys.argv[1] if len(sys.argv) > 1 else 'pricing-engine-v2-input.xlsx'
    print('  wrote ' + build(out))
