#!/usr/bin/env python3
"""Assert that the pricing workbook contains no business values.

The workbook's whole purpose is to be empty: it asks the business for rates and
must never hand back a number that looks like an answer. This checks that every
sheet is present, that every populated data row is an example, that every example
row says so, and that every numeric value on those rows is zero.

    python3 tests/tools/check-pricing-workbook.py <workbook.xlsx>
"""
import sys
from openpyxl import load_workbook

SHEETS = ['README', 'MATERIAL_RATES', 'DIAMETER_RULES', 'PROCESS_COST_RULES',
          'LABOUR_QTY_RULES', 'FINISH_COST_RULES', 'CUSTOMER_MARKUP', 'ACCESSORY_COSTS']

asserts = 0
failures = []


def ok(cond, msg):
    global asserts
    asserts += 1
    if not cond:
        failures.append(msg)


def main(path):
    wb = load_workbook(path)
    ok(wb.sheetnames == SHEETS, 'the workbook has exactly the sheets asked for: %s' % (wb.sheetnames,))

    readme = '\n'.join(str(c.value or '') for c in wb['README']['A'])
    ok('EXAMPLE — DELETE' in readme, 'the README tells the reader to delete the example rows')
    ok('Effective Date' in readme, 'and explains effective dates')
    ok('Cost Basis' in readme, 'and explains the cost basis')

    for name in SHEETS[1:]:
        ws = wb[name]
        header_row = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            # the table header repeats the first column name with a real second column
            if a and b and b not in ('REQUIRED', 'optional') and header_row is None and r > 5:
                header_row = r
                break
        ok(header_row is not None, '%s: has a header row' % name)
        if header_row is None:
            continue

        headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        ok(all(h for h in headers), '%s: every column is named' % name)
        ok('Remark' in headers, '%s: has a Remark column' % name)
        ok('Active' in headers, '%s: has an Active column' % name)

        # Only the columns that carry a business NUMBER. "Cost Basis" is a unit
        # label and "Min Value" is a condition, not a price.
        MONEY = {'Supplier Cost Per KG', 'Internal Cost Rate Per KG', 'Diameter (mm)',
                 'Cost', 'Labour Cost Per PC', 'Markup %', 'Supplier Cost'}
        money = [i for i, h in enumerate(headers) if h in MONEY]
        remark_at = headers.index('Remark')
        active_at = headers.index('Active')

        populated = 0
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c + 1).value for c in range(len(headers))]
            if not any(v not in (None, '') for v in row):
                continue
            populated += 1
            remark = str(row[remark_at] or '')
            ok('EXAMPLE — DELETE' in remark,
               '%s row %d: a populated row is marked as an example (%r)' % (name, r, remark[:60]))
            ok(str(row[active_at] or '').upper() == 'N',
               '%s row %d: an example row is inactive' % (name, r))
            for i in money:
                v = row[i]
                ok(v in (None, '', 0, 0.0),
                   '%s row %d: %s carries no invented value (found %r)' % (name, r, headers[i], v))
        ok(populated <= 1, '%s: at most one example row, so the shape is shown and nothing else' % name)

    print('  %s  pricing workbook — structure present, no business values  (%d assertions%s)'
          % ('FAIL' if failures else 'ok  ', asserts,
             ', %d failed' % len(failures) if failures else ''))
    for f in failures:
        print('   - ' + f)
    return 1 if failures else 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else 'pricing-engine-v2-input.xlsx'))
